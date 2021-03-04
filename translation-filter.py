from openpyxl import load_workbook
import argparse
import re
import logging


def createlogger(name):
    """Create a logger named specified name with the level set in config file.
    """
    logger = logging.getLogger(name)
    logger.setLevel("DEBUG")
    if logger.handlers:
        for h in logger.handlers:
            logger.removeHandler(h)
    ch = logging.StreamHandler()
    formatter = logging.Formatter(
        '%(asctime)s.%(msecs)03d: [%(levelname)s] [%(name)s] [%(funcName)s] %(message)s',
        '%y%m%d %H:%M:%S')
    ch.setFormatter(formatter)
    logger.addHandler(ch)
    return logger


logger = createlogger("MAIN")


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("-f", "--file", required=True, help='excel file with path')
    parser.add_argument("-r", "--replace", action="store_true", default=False, help='replace old excel')
    parser.add_argument("-c", "--column", required=False, default=4, help='Text column number')
    return parser.parse_args()


def re_match(string: str):
    result = re.sub(r'<.*?>', '', string)
    return result


# draft
def singel_cell(old_cell_value):
    cnt_list = old_cell_value.split("</xliff:g>")
    new_cell_value = ""
    # 正则匹配方式
    # print(f"cnt_list：{cnt_list}")
    # for i in cnt_list:
    #     if "<xliff:g" in i:
    #         result_list = re_match(i)
    #         print(f"result_list:{result_list}")
    #         result_str = ""
    #         for result in result_list:
    #             result_str += result[0] + result[1]
    #         new_cell_value += str(result_str)
    #     else:
    #         new_cell_value += i
    # return new_cell_value

    # 字符串
    # for i in cnt_list:
    #     if "<xliff" in i:
    #         for j in i.split("<"):
    #             if ">" in j:
    #                 new_cell_value += j.split(">")[1]
    #             else:
    #                 new_cell_value += j
    #     else:
    #         new_cell_value += i
    return new_cell_value


def load_excel_file(file, column, replace):
    wb = load_workbook(file)
    sheet = wb.worksheets[0]
    max_raw = sheet.max_row

    for i in range(1, max_raw):
        cell_cnt = sheet.cell(i, int(column)).value
        if cell_cnt and "<" in cell_cnt:
            # single cell test
            # if "Only one downloaded SIM" in cell_cnt:
            #     print(i)
            #     print(cell_cnt)
            #     new_cell = singel_cell(cell_cnt)
            #     print(new_cell)
            logger.info(f"row:{i}")
            logger.info(f"old cell:{cell_cnt}")
            new_cell = re_match(cell_cnt)
            logger.info(f"new cell:{new_cell}")
            sheet.cell(i, int(column), new_cell)
    if replace:
        file_name = file
    else:
        file_name = file.replace(".xlsx", "_new.xlsx")
    wb.save(file_name)
    return file_name


if __name__ == '__main__':
    args = parse_args()
    file_name = load_excel_file(args.file, args.column, args.replace)
    logger.info(f"data already save to {file_name}")
